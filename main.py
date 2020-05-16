import openpyxl
import os
from openpyxl.drawing.image import Image

from config import Config
from libs import *

print("Phase analysis program. Sponsored by NATO & NASA")

workbook = openpyxl.load_workbook(Config.workbook)
if Config.sheet != '':
    source_sheet = workbook[Config.sheet]
else:
    source_sheet = workbook[workbook.sheetnames[0]]

# импортируем данные на новые лист
import_source_data(workbook, source_sheet, Config)

sheet = workbook[Config.output_sheet]

# считаем приращение и сдвиг
calculate_derivative(sheet, Config)

os.remove(Config.workbook)
workbook.save(Config.workbook)


# находим квазициклы
quasicycles = get_quasicycles(sheet, Config)

# рисуем квазициклы
col_index = 0
row_index = 0
for quasicycle in quasicycles:
    sheet.add_chart(create_diagram(quasicycle), str(sheet.cell(row_index * 15 + 1, 5 + (col_index % 2) * 6).coordinate))
    if col_index % 2:
        row_index += 1
    col_index += 1

q_size = Config.start_row
while sheet.cell(row=q_size + 1, column=3).value is not None:
    q_size += 1
q_size -= Config.start_row

# рисуем фазовый портрет
phase_portrait = Quasicycle(sheet, "Фазовый портрет", Config.start_row, 2, q_size)
sheet.add_chart(create_diagram(phase_portrait, 20, 20, 12), str(sheet.cell(1, 17).coordinate))

# память квазициклов
index = 1
current_row = q_size + 9 + index
sheet.cell(current_row, 1).value = "Номер квазицикла"
sheet.cell(current_row, 2).value = "Память квазицикла"
current_row += 1
start_row = current_row
for quasicycle in quasicycles:
    sheet.cell(current_row, 1).value = index
    sheet.cell(current_row, 2).value = quasicycle.size
    index += 1
    current_row += 1
sheet.add_chart(create_bar_chart(sheet, start_row, index - 2), str(sheet.cell(40, 17).coordinate))

# движения площадей прямоугольников
create_squares_graph(quasicycles)
img = Image('squares.png')
sheet.add_image(img, str(sheet.cell(56, 17).coordinate))

os.remove(Config.workbook)
workbook.save(Config.workbook)
