from openpyxl.chart import Reference, ScatterChart, LineChart, BarChart, Series
from quasicycle import Quasicycle
import numpy as np
from sklearn.cluster import MeanShift, KMeans
import matplotlib.pyplot as plt


def import_source_data(workbook, source_sheet, config):
    col = 1
    while True:
        if source_sheet.cell(row=1, column=col).value == config.source_col:
            break
        col += 1
    if config.output_sheet in workbook.sheetnames:
        workbook.remove_sheet(workbook.get_sheet_by_name(config.output_sheet))
    workbook.create_sheet(config.output_sheet)
    output_sheet = workbook[config.output_sheet]
    i = config.start_row
    while source_sheet.cell(row=i + 1, column=col).value is not None:
        output_sheet.cell(i, 1).value = float(source_sheet.cell(i + 1, col).value)
        i += 1


def create_diagram(quasicycle, height=7, width=10, style=11):
    chart = ScatterChart()
    chart.title = quasicycle.name
    chart.height = height
    chart.width = width
    chart.x_axis.title = ''
    chart.y_axis.title = ''
    chart.legend = None
    rows_reference = Reference(quasicycle.sheet, min_col=quasicycle.start_cell_col,
                               min_row=quasicycle.start_cell_row, max_row=quasicycle.start_cell_row + quasicycle.size)
    cols_reference = Reference(quasicycle.sheet, min_col=quasicycle.start_cell_col + 1,
                               min_row=quasicycle.start_cell_row, max_row=quasicycle.start_cell_row + quasicycle.size)
    series = Series(cols_reference, rows_reference, title_from_data=False)
    chart.layoutTarget = "inner"
    chart.style = style
    chart.series.append(series)
    return chart


def create_bar_chart(sheet, start_row, size):
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Память квазициклов"
    bar_chart.y_axis.title = 'Память'
    bar_chart.x_axis.title = 'Квазициклы'
    data = Reference(sheet, min_col=2, min_row=start_row - 1, max_row=start_row + size)
    indexes = Reference(sheet, min_col=1, min_row=start_row, max_row=start_row + size)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(indexes)
    bar_chart.shape = 4
    bar_chart.legend = None
    return bar_chart


def calculate_derivative(sheet, config):
    sheet.cell(1, 1).value = "Временной ряд"
    sheet.cell(1, 2).value = "Приращение"
    sheet.cell(1, 3).value = "Сдвиг"
    current_row = config.start_row
    while sheet.cell(current_row + 1, config.start_col).value is not None:
        sheet.cell(current_row, config.start_col + 1).value = sheet.cell(current_row + 1, config.start_col).value \
                                                              - sheet.cell(current_row, config.start_col).value
        current_row += 1
    current_row = config.start_row
    while sheet.cell(current_row + 1, config.start_col + 1).value is not None:
        sheet.cell(current_row, config.start_col + 2).value = sheet.cell(current_row + 1, config.start_col + 1).value
        current_row += 1


def distance(point1, point2):
    square_x = (point2[0] - point1[0]) ** 2
    square_y = (point2[1] - point1[1]) ** 2
    dist = (square_x + square_y) ** 0.5
    return dist


def check_near(points_list, time_check, position, next_num, min_value):
    while time_check != 0:
        try:
            if distance(points_list[position], points_list[position + next_num + 1]) < min_value:
                min_value = distance(points_list[position], points_list[position + next_num + 1])
                next_num += 1
            time_check -= 1
        except IndexError:
            break
    return min_value, next_num


def get_quasicycles(sheet, config):
    quasicycles = []
    points_list = []
    row = config.start_row
    while sheet.cell(row, 3).value is not None:
        points_list.append([sheet.cell(row, 2).value, sheet.cell(row, 3).value])
        row += 1
    position = 0
    q_index = 1
    while position + config.min_size < len(points_list) - 1:
        q_size = config.min_size
        min_value = distance(points_list[position], points_list[q_size + position])
        time_check = 3
        min_value, q_size = check_near(points_list, time_check, position, q_size, min_value)
        if q_size == q_size + time_check:
            while distance(points_list[position], points_list[position + q_size + 1]) < min_value:
                min_value = distance(points_list[position], points_list[position + q_size])
                q_size += 1
        quasicycles.append(Quasicycle(sheet, "Квазицикл " + str(q_index), position + config.start_row, 2, q_size))
        position = position + q_size + 1
        q_index += 1
    return quasicycles


def create_squares_graph(quasicycles):
    sort_quasi = []
    order = 1
    for quasicycle in quasicycles:
        sort_quasi.append([order, quasicycle.square])
        order += 1
    sort_quasi = np.array(sort_quasi)
    sort_quasi.reshape(2, order - 1)
    k_means = KMeans(n_clusters=3)
    k_means.fit(sort_quasi)
    y_k_means = k_means.predict(sort_quasi)
    fig, ax = plt.subplots()
    ax.scatter(sort_quasi[:, 0], sort_quasi[:, 1], c=y_k_means, s=50, cmap='viridis')
    plt.title('График движений площадей прямоугольников')
    plt.xlabel('Номер квазицикла')
    plt.ylabel('Площадь квазицикла')
    fig.savefig('squares.png')
    # plt.show()
