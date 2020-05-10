import openpyxl

Datwb = openpyxl.load_workbook(filename='Data/BaseData.xlsx')
CalcSheet = Datwb['BaseData']

#ищем нужный стобец
col=1
while True:
    if CalcSheet.cell(row=1, column=col).value == "<CLOSE>":
        break
    col+=1

#ищем нужный диапазон
rw=2
while True:
    if CalcSheet.cell(row=rw, column=col).value ==  None:
        break
    rw += 1
rw-=1

#находим координаты начала и конца столбца
xy1 = str(CalcSheet.cell(1,col).coordinate)
xy2 = str(CalcSheet.cell(rw,col).coordinate)

#двигаем колонку данных в начало
CalcSheet.move_range(xy1+":"+xy2, rows=0, cols=-col+1)

#удаляем все остальные колонки
emp = 2
while True:
    if CalcSheet.cell(1,emp).value == None:
        break
    emp+=1
CalcSheet.delete_cols(2, amount=emp)


#приводим к типу float, удаляем delimeter
ctrrw = 2
while CalcSheet.cell(row=ctrrw, column=1).value !=  None:
      if type(CalcSheet.cell(ctrrw, 1).value) is str:
        CalcSheet.cell(ctrrw, 1).value = (CalcSheet.cell(ctrrw, 1).value).replace(",", "")
        CalcSheet.cell(ctrrw, 1).value = float(CalcSheet.cell(ctrrw, 1).value)
      else:
        CalcSheet.cell(ctrrw, 1).value = float(CalcSheet.cell(ctrrw, 1).value)
      ctrrw+=1

#считаем приращение
CalcSheet.cell(1,2).value = "Приращение"
iter=2
while CalcSheet.cell(iter+1,1).value != None:
   CalcSheet.cell(iter, 2).value = CalcSheet.cell(iter+1, 1).value - CalcSheet.cell(iter, 1).value
   iter += 1

#считаем сдвиг
iter=2
CalcSheet.cell(1,3).value = "Сдвиг"
while CalcSheet.cell(iter+1,2).value!= None:
    CalcSheet.cell(iter, 3).value = CalcSheet.cell(iter + 1, 2).value
    iter+=1

Datwb.save('Data/BaseData.xlsx')
